import typer
from logging import getLogger
from util.logger import setup_root_logger
import openpyxl
import re
import os
import glob
import win32com.client
from datetime import datetime, timedelta
import holidays  # 日本の祝日データを取得するライブラリ (install: `pip install holidays`)

logger = getLogger()
app = typer.Typer()

def generate_expected_sheet_names(start_yyyymm: str, end_yyyymm: str):
    """指定された年月の範囲で、各週の月曜始まり・日曜終わりのシート名リストを生成"""
    expected_sheets = []
    
    start_date = datetime.strptime(start_yyyymm, "%Y%m").replace(day=1)
    end_date = datetime.strptime(end_yyyymm, "%Y%m").replace(day=1)
    
    next_month = end_date.month % 12 + 1
    next_month_year = end_date.year + (1 if next_month == 1 else 0)
    last_day = datetime(next_month_year, next_month, 1) - timedelta(days=1)
    
    current_date = start_date

    while current_date <= last_day:
        if current_date.weekday() != 0:
            current_date += timedelta(days=(7 - current_date.weekday()))
        
        if current_date > last_day:
            break
        
        sunday = current_date + timedelta(days=6)
        expected_sheets.append(f"{current_date.strftime('%Y%m%d')}_{sunday.strftime('%Y%m%d')}")
        
        current_date += timedelta(days=7)

    return expected_sheets

def check_sheet_dates(ws, sheet_name):
    """指定シートの H10 〜 H16 に正しい日付・数式が入っているか確認"""
    match = re.match(r"^(\d{8})_(\d{8})$", sheet_name)
    if not match:
        logger.warning(f"シート名が不正です: {sheet_name}")
        return False

    start_date = datetime.strptime(match.group(1), "%Y%m%d")

    # H10 の日付チェック
    h10_value = ws["H10"].value
    if isinstance(h10_value, datetime):
        h10_date = h10_value
    else:
        try:
            h10_date = datetime.strptime(str(h10_value), "%Y/%m/%d")
        except ValueError:
            logger.warning(f"シート {sheet_name} の H10 の値が不正: {h10_value}")
            return False

    if h10_date != start_date:
        logger.warning(f"シート {sheet_name} の H10 の値が正しくありません: {h10_date} (期待値: {start_date.strftime('%Y/%m/%d')})")
        return False

    incorrect_cells = []

   # H11～H16 の数式チェック
    for i in range(1, 7):
        cell_ref = f"H{10 + i}"
        cell_value = ws[cell_ref].value
        expected_formula = f"=H{9 + i}+1"

        if not isinstance(cell_value, str) or not cell_value.startswith("="):
            incorrect_cells.append(f"{cell_ref}: 数式が設定されていません (期待値: {expected_formula})")
        elif cell_value != expected_formula:
            incorrect_cells.append(f"{cell_ref}: {cell_value} (期待値: {expected_formula})")

    # A列の数式チェック (H10 ～ H16 を参照)
    for i, h_row in enumerate(range(10, 17)):
        a_row = 10 + (i * 6)
        h_ref = f"H{h_row}"

        expected_month_formula = f"=MONTH({h_ref})"
        expected_day_formula = f"=DAY({h_ref})"
        expected_text_formula = f'="("&TEXT({h_ref}, "aaa")&")"'

        if ws[f"A{a_row}"].value != expected_month_formula:
            incorrect_cells.append(f"A{a_row}: 数式が正しくありません (期待値: {expected_month_formula})")

        if ws[f"A{a_row + 1}"].value != expected_day_formula:
            incorrect_cells.append(f"A{a_row + 1}: 数式が正しくありません (期待値: {expected_day_formula})")

        if ws[f"A{a_row + 2}"].value != expected_text_formula:
            incorrect_cells.append(f"A{a_row + 2}: 数式が正しくありません (期待値: {expected_text_formula})")

    # B列の固定値チェック (指定セル範囲で完全一致)
    expected_b_values = {
        10: "月", 11: "日",
        16: "月", 17: "日",
        22: "月", 23: "日",
        28: "月", 29: "日",
        34: "月", 35: "日",
        40: "月", 41: "日",
        46: "月", 47: "日",
    }

    for row, expected_value in expected_b_values.items():
        cell_ref = f"B{row}"
        actual_value = ws[cell_ref].value

        if actual_value != expected_value:
            incorrect_cells.append(f"{cell_ref}: 値が正しくありません (期待値: '{expected_value}', 実際: '{actual_value}')")

    if incorrect_cells:
        logger.warning(f"シート {sheet_name} のセル値が正しくありません:\n" + "\n".join(incorrect_cells))
        return False

    return True



def check_daily_report(ws, sheet_name):
    """ C列(業務内容)に未入力があるかチェック"""
    required_cells = ["C9", "C15", "C21", "C27", "C33", "C39", "C45"]
    empty_cells = [cell for cell in required_cells if not ws[cell].value]
    if empty_cells:
        logger.warning(f"シート {sheet_name} の日報が入力されていません: {', '.join(empty_cells)}")
        return False
    # C39, C45 の値が "休暇" であることをチェック
    for cell in ["C39", "C45"]:
        if ws[cell].value != "休暇":
            logger.warning(f"シート {sheet_name} の {cell} の値が '休暇' ではありません (実際: {ws[cell].value})")
            return False
    return True

import holidays

import holidays

import holidays
import logging
from datetime import datetime

logger = logging.getLogger(__name__)
import logging

logger = logging.getLogger(__name__)

import logging

logger = logging.getLogger(__name__)

def check_specific_entries(ws, sheet_name):
    """
    - A57 に `=H14` という数式が入力されているかチェック
    - F4 に `miracleave株式会社` という文字列が入力されているかチェック
    - C6 が未入力でないかチェック（数値・文字列含む）
    """
    errors = []

    # A57セルのチェック
    a57_formula = ws["A57"].value
    if a57_formula != "=H14":
        errors.append(f"A57: {a57_formula} ('=H14' ではありません)")

    # F4セルのチェック
    f4_value = ws["F4"].value
    if f4_value != "miracleave株式会社":
        errors.append(f"F4: {f4_value} ('miracleave株式会社' ではありません)")

    # C6セルのチェック（数値や文字列を含め、未入力とする）
    c6_value = ws["C6"].value
    if c6_value is None or (isinstance(c6_value, str) and c6_value.strip() == ""):
        errors.append("C6: 未入力です")

    # エラーがあればログに出力
    if errors:
        logger.warning(f"シート {sheet_name} のチェックに失敗しました:\n" + "\n".join(errors))
        return False

    return True



def check_holiday_entries(ws, sheet_name):
    """H10〜H16の日付が祝日または土日の場合、C列に「祝日」または「休暇」が入力されているかチェック。
       平日の場合、「祝日」または「休暇」が入力されていないかチェック。"""
    jp_holidays = holidays.JP()  # 日本の祝日情報を取得

    # H列とC列の対応表
    check_map = {
        "H10": "C9",
        "H11": "C15",
        "H12": "C21",
        "H13": "C27",
        "H14": "C33",
        "H15": "C39",
        "H16": "C45",
    }

    incorrect_entries = []

    # H10 の日付を取得
    h10_value = ws["H10"].value
    if isinstance(h10_value, datetime):
        h10_date = h10_value  # すでに datetime 型ならそのまま
    else:
        try:
            h10_date = datetime.strptime(str(h10_value), "%Y-%m-%d %H:%M:%S")  # Excel のフォーマットに合わせる
        except ValueError:
            logger.warning(f"シート {sheet_name} の H10 の値が不正: {h10_value}")
            return False  # H10 が不正なら処理を中断

    for h_cell, c_cell in check_map.items():
        if h_cell == "H10":
            h_date = h10_date  # H10 はそのまま使用
        else:
            day_offset = int(h_cell[1:]) - 10  # H11 → +1, H12 → +2, …
            h_date = h10_date + timedelta(days=day_offset)

        c_value = ws[c_cell].value

        # 祝日 or 土日チェック
        is_holiday_or_weekend = (h_date in jp_holidays) or (h_date.weekday() >= 5)

        if is_holiday_or_weekend:
            if c_value not in ["祝日", "休暇"]:
                incorrect_entries.append(f"{c_cell}: {c_value} (祝日・週末なのに '祝日' または '休暇' ではありません)")
        else:
            if c_value in ["祝日", "休暇"]:
                incorrect_entries.append(f"{c_cell}: {c_value} (平日なのに '祝日' または '休暇' が入力されています)")

    if incorrect_entries:
        logger.warning(f"シート {sheet_name} の日付エントリが不正です:\n" + "\n".join(incorrect_entries))
        return False

    return True






@app.command("check")
def sheet_name_check(start_yyyymm: str, end_yyyymm: str):
    input_dir = os.path.abspath("input")
    excel_files = glob.glob(os.path.join(input_dir, "*.xlsx"))
    if not excel_files:
        raise FileNotFoundError("input フォルダに Excel ファイルが見つかりません。")
    input_path = excel_files[0]
    wb = openpyxl.load_workbook(f"{input_path}", data_only=False)
    existing_sheets = {ws.title for ws in wb.worksheets}

    expected_sheets = generate_expected_sheet_names(start_yyyymm, end_yyyymm)

    missing_sheets = [sheet for sheet in expected_sheets if sheet not in existing_sheets]
    extra_sheets = [sheet for sheet in existing_sheets if not re.match(r'^\d{8}_\d{8}$', sheet)]

    if missing_sheets:
        logger.warning(f"不足しているシートがあります: {missing_sheets}")
    else:
        logger.info("必要なシートはすべて存在しています。")

    if extra_sheets:
        logger.warning(f"適切ではないシート名が検出されました: {extra_sheets}")

    for sheet_name in expected_sheets:
        if sheet_name in existing_sheets:
            ws = wb[sheet_name]
            check_sheet_dates(ws, sheet_name)
            check_daily_report(ws, sheet_name)
            check_holiday_entries(ws, sheet_name)
            check_specific_entries(ws, sheet_name)

@app.command("cut")
def cut_out_sheet(start_yyyymm: str, end_yyyymm: str):
    """
    指定した年月範囲に含まれるシートのみを抽出し、新しいExcelファイルとして出力する。
    """
    input_dir = os.path.abspath("input")
    # inputフォルダ内の最初に見つかった.xlsxファイルを取得
    excel_files = glob.glob(os.path.join(input_dir, "*.xlsx"))
    if not excel_files:
     raise FileNotFoundError("input フォルダに Excel ファイルが見つかりません。")

    input_path = excel_files[0]  # 最初のExcelファイルを取得
    output_dir = os.path.abspath("output")
    output_filename = os.path.join(output_dir, f"日報_抽出_{start_yyyymm}_{end_yyyymm}.xlsx")
    
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    try:
        # Excelアプリケーションの起動
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excelを非表示で実行

        # 元のExcelファイルを開く
        wb = excel.Workbooks.Open(input_path)
        
        # 新規Excelファイルを作成
        new_wb = excel.Workbooks.Add()

        # 期待されるシート名を取得（順番通り）
        expected_sheets = generate_expected_sheet_names(start_yyyymm, end_yyyymm)

        # シートを順番通りにコピー
        copied_any = False
        for sheet_name in expected_sheets:
            if sheet_name in [sheet.Name for sheet in wb.Sheets]:
                wb.Sheets(sheet_name).Copy(Before=new_wb.Sheets(1))
                copied_any = True

        if copied_any:
            # 最初に自動作成された空のシートがある場合は削除
            if new_wb.Sheets.Count > 1:
                new_wb.Sheets(new_wb.Sheets.Count).Delete()

            # 新規Excelファイルを保存
            new_wb.SaveAs(output_filename)
            logger.info(f"抽出したシートを {output_filename} に出力しました。")
        else:
            logger.warning("指定範囲に該当するシートがありませんでした。")

        # ファイルを閉じる
        new_wb.Close(SaveChanges=False)
        wb.Close(SaveChanges=False)

    except Exception as e:
        logger.error(f"エラーが発生しました: {e}")
    
    finally:
        excel.Quit()  # Excelを終了

import os
import glob
import win32com.client
import typer
from logging import getLogger

@app.command("move_a1")
def move_active_cell_to_a1():
    """
    全てのシートのアクティブセルを A1 に移動し、最初のシートをアクティブにして上書き保存する
    """
    input_dir = os.path.abspath("input")
    excel_files = glob.glob(os.path.join(input_dir, "*.xlsx"))

    if not excel_files:
        raise FileNotFoundError("input フォルダに Excel ファイルが見つかりません。")

    input_path = excel_files[0]  # 最初に見つかったExcelファイルを処理
    logger.info(f"処理対象のExcelファイル: {input_path}")

    try:
        # Excelアプリケーションを起動
        excel = win32com.client.Dispatch("Excel.Application")
        excel.Visible = False  # Excelを非表示で実行

        # 指定のExcelファイルを開く
        wb = excel.Workbooks.Open(input_path)

        # すべてのシートのアクティブセルをA1に設定
        for sheet in wb.Sheets:
            sheet.Activate()
            sheet.Range("A1").Select()

        # **最初のシートをアクティブにする**
        first_sheet = wb.Sheets(1)  # 1番目のシートを取得
        first_sheet.Activate()  # アクティブに設定

        # 上書き保存
        wb.Save()
        logger.info(f"すべてのシートのアクティブセルをA1に移動し、最初のシートをアクティブにして保存しました: {input_path}")

        # ファイルを閉じる
        wb.Close(SaveChanges=True)

    except Exception as e:
        logger.error(f"エラーが発生しました: {e}")

    finally:
        excel.Quit()  # Excelアプリケーションを終了



if __name__ == "__main__":
    setup_root_logger(verbose=True)
    app()
