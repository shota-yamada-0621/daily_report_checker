import typer
from logging import getLogger
from util.logger import setup_root_logger
import openpyxl
import re
import os
import glob
import win32com.client
from datetime import datetime, timedelta
import holidays  # 日本の祝日データを取得するライブラリ (install: `pip install holidays`)

logger = getLogger()
app = typer.Typer()

def generate_expected_sheet_names(start_yyyymm: str, end_yyyymm: str):
    """指定された年月の範囲で、各週の月曜始まり・日曜終わりのシート名リストを生成"""
    expected_sheets = []
    
    start_date = datetime.strptime(start_yyyymm, "%Y%m").replace(day=1)
    end_date = datetime.strptime(end_yyyymm, "%Y%m").replace(day=1)
    
    next_month = end_date.month % 12 + 1
    next_month_year = end_date.year + (1 if next_month == 1 else 0)
    last_day = datetime(next_month_year, next_month, 1) - timedelta(days=1)
    
    current_date = start_date

    while current_date <= last_day:
        if current_date.weekday() != 0:
            current_date += timedelta(days=(7 - current_date.weekday()))
        
        if current_date > last_day:
            break
        
        sunday = current_date + timedelta(days=6)
        expected_sheets.append(f"{current_date.strftime('%Y%m%d')}_{sunday.strftime('%Y%m%d')}")
        
        current_date += timedelta(days=7)

    return expected_sheets

def check_sheet_dates(ws, sheet_name):
    """指定シートの H10 〜 H16 に正しい日付・数式が入っているか確認"""
    match = re.match(r"^(\d{8})_(\d{8})$", sheet_name)
    if not match:
        logger.warning(f"シート名が不正です: {sheet_name}")
        return False

    start_date = datetime.strptime(match.group(1), "%Y%m%d")

    # H10 の日付チェック
    h10_value = ws["H10"].value or ws["H10"].internal_value
    if isinstance(h10_value, datetime):
        h10_date = h10_value
    else:
        try:
            h10_date = datetime.strptime(str(h10_value), "%Y/%m/%d")
        except ValueError:
            logger.warning(f"シート {sheet_name} の H10 の値が不正: {h10_value}")
            return False

    if h10_date != start_date:
        logger.warning(f"シート {sheet_name} の H10 の値が正しくありません: {h10_date} (期待値: {start_date.strftime('%Y/%m/%d')})")
        return False

    return True

def check_daily_report(ws, sheet_name):
    """ C列(業務内容)に未入力があるかチェック"""
    required_cells = ["C9", "C15", "C21", "C27", "C33", "C39", "C45"]
    empty_cells = [cell for cell in required_cells if not ws[cell].value]
    if empty_cells:
        logger.warning(f"シート {sheet_name} の日報が入力されていません: {', '.join(empty_cells)}")
        return False
    
    # C39, C45 の値が "休暇" であることをチェック
    for cell in ["C39", "C45"]:
        if ws[cell].value != "休暇":
            logger.warning(f"シート {sheet_name} の {cell} の値が '休暇' ではありません (実際: {ws[cell].value})")
            return False
    
    return True

@app.command("check")
def sheet_name_check(start_yyyymm: str, end_yyyymm: str):
    input_dir = os.path.abspath("input")
    excel_files = glob.glob(os.path.join(input_dir, "*.xlsx"))
    if not excel_files:
        raise FileNotFoundError("input フォルダに Excel ファイルが見つかりません。")
    input_path = excel_files[0]
    wb = openpyxl.load_workbook(f"{input_path}", data_only=False)
    existing_sheets = {ws.title for ws in wb.worksheets}

    expected_sheets = generate_expected_sheet_names(start_yyyymm, end_yyyymm)

    missing_sheets = [sheet for sheet in expected_sheets if sheet not in existing_sheets]
    extra_sheets = [sheet for sheet in existing_sheets if not re.match(r'^\d{8}_\d{8}$', sheet)]

    if missing_sheets:
        logger.warning(f"不足しているシートがあります: {missing_sheets}")
    else:
        logger.info("必要なシートはすべて存在しています。")

    if extra_sheets:
        logger.warning(f"適切ではないシート名が検出されました: {extra_sheets}")

    for sheet_name in expected_sheets:
        if sheet_name in existing_sheets:
            ws = wb[sheet_name]
            check_sheet_dates(ws, sheet_name)
            check_daily_report(ws, sheet_name)

if __name__ == "__main__":
    setup_root_logger(verbose=True)
    app()
